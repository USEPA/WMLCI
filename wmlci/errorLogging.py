import pandas as pd
from collections import defaultdict
import os
import hashlib
from typing import Optional, List
from bw2calc import LCA, LeastSquaresLCA
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import zipfile

from bw2io.importers.json_ld import JSONLDImporter

from wmlci.settings import paths, sourcedatapath
from wmlci.wmlci_log import log
from esupy.remote import make_url_request
from esupy.util import make_uuid
from esupy.processed_data_mgmt import download_from_remote, Paths, mkdir_if_missing
"""
Functions for locating where incompatibilities exist between olca json-ld and brightway.
"""
def print_avoided_input_uuids(jsonld):
    """
    Prints the UUIDs of processes and exchanges where avoided products are used as inputs.

    Parameters:
        jsonld (JSONLDImporter): An initialized JSONLDImporter object with data loaded.
    """
    log.info("\nChecking for avoided products used as inputs\n")
    for pid, process in jsonld.data.get("processes", {}).items():
        if process.get("input"):
            for exc in process.get("exchanges", []):
                if exc.get("isAvoidedProduct"):
                    log.info(f" Process UUID: {pid} -> Exchange UUID: {exc.get('id')}")
    log.info("\n✅ Scan complete.")

def find_missing_unit_group_id(ug_id, jsonld):
    """
    search for missing unit group uuid(s) in importer object
    print process and exchange info for missing unit group id
    :param ug_id:
    :param jsonld:
    :return:
    """
    print("🔍 Scanning processes for unit group issues...\n")
    for pid, process in jsonld.data.get("processes", {}).items():
        for exc in process.get("exchanges", []):
            unit = exc.get("unit", {})
            if isinstance(unit, dict) and unit.get("@id") == ug_id:
                print(f"\n⚠️ Problem in activity: \n--Process: {pid}; Exchange: {exc}")
                print("-" * 60)
    return "\n✅ Search for unit group id issues is complete."

def find_production_exchange_errors(jsonld):
    """
    search for processes with zero or more than one exchanges of flowType PRODUCT_FLOW
    print process and exchange info for missing unit group id
    this is a debugging function that should help find causes of the 'Failed Allocation' assertion error

    running methods edit_non_quant_ref_flow_type() and apply_opposite_direction_approach() should result
    in this method producing no outputs

    :param jsonld:
    :return:
    """
    print("🔍 Scanning processes for production exchange issues...\n")
    for process_id, process in jsonld.data.get("processes", {}).items():
        if process.get("type") in {"emission", "product"}:
            continue
        exchanges = process.get("exchanges", [])
        production_exchanges = [
            exc for exc in exchanges
            if exc.get("flow", {}).get("flowType") == "PRODUCT_FLOW" and not exc.get("isInput")
        ]
        if len(production_exchanges) != 1:
            print(f"⚠️  Problem in activity: {process.get('name', 'Unnamed')} (ID: {process_id})")
            print(f"Found {len(production_exchanges)} production exchanges:")
            for exc in production_exchanges:
                flow_name = exc.get("flow", {}).get("name", "Unknown")
                is_input = exc.get("isInput")
                print(f"  - Flow: {flow_name}, isInput: {is_input}")
            print("-" * 60)
        else:
            process["unit"] = production_exchanges[0]["unit"]
    return "\n✅ Search for production exchange issues is complete."

def find_location_issues(jsonld):
    """
    search for processes and exchanges with missing location info
    print process and exchange info where location info is missing
    this is a debugging function that should help find causes 'location' key errors

    running methods append_jsonld_location(jsonld) should result in no outputs

    :param jsonld:
    :return:
    """
    print("🔍 Scanning processes for location issues...\n")
    for process_id, process in jsonld.data.get("processes", {}).items():
        if process.get("type") in {"emission", "product"}:
            continue
        location = process.get("location")
        if isinstance(location, dict) and "name" in location:
            continue
        else:
            print(f"⚠️ Problem with location in process ID '{process_id}':")
            print(f"  Name: {process.get('name', 'Unnamed')}")
            print(f"  Location value: {location}")
            print("-" * 60)
    return "\n✅ Search for location issues is complete."

def find_faulty_allocation_factors(jsonld):
    """
    used to solve: "We currently only support exchange-specific CAUSAL_ALLOCATION"

    search for processes with allocationFactors that are missing  'exchange', 'product', or both
    print the process uuid and other useful information

    deleting entries in allocationFactors which has the allocation type of 'CAUSAL_ALLOCATION' but no exchange attribute
    will solve this problem.

    :param jsonld:
    :return:
    """
    faulty_processes = []

    for process_id, process in jsonld.data.get("processes", {}).items():
        if process.get("type") in {"emission", "product"}:
            continue

        for idx, factor in enumerate(process.get("allocationFactors", [])):
            allocation_type = factor.get("allocationType", "UNKNOWN_ALLOCATION")
            missing = []

            if "product" not in factor:
                missing.append("product")
            if allocation_type == "CAUSAL_ALLOCATION" and "exchange" not in factor:
                missing.append("exchange")

                faulty_processes.append({
                    "process_id": process_id,
                    "allocation_type": allocation_type,
                    "missing_components": missing,
                    "factor_index": idx
                })

    print("📋 Summary of faulty processes:")
    for entry in faulty_processes:
        print(
            f"- Process ID: {entry['process_id']}, "
            f"Allocation Type: {entry['allocation_type']}, "
            f"Missing: {', '.join(entry['missing_components'])}, "
            f"Factor Index: {entry['factor_index']}"
        )

    return faulty_processes

def find_unallocatable_processes(jsonld):
    """
    used to solve: "Default allocation chosen, but allocation factors for this method not provided"

    Identify and print processes in a JSONLDImporter object that require allocation
    but lack allocation factors for their default allocation method. This also occurs when the default allocation is set
    to 'NO_ALLOCATION" but allocation is required.

    changing the default allocation to an allocation type which match allocation factors included in attribute
    'allocationFactors'.

    :param jsonld:
    :return:
    """
    for process_id, process in jsonld.data.get("processes", {}).items():
        # Skip if not allocatable
        if process.get("@type") in ("product", "emission"):
            continue
        if not process.get("allocationFactors"):
            continue
        allocation_dict = defaultdict(dict)
        try:
            for factor in process["allocationFactors"]:
                allocation_type = factor.get("allocationType")
                if allocation_type == "CAUSAL_ALLOCATION":
                    try:
                        product = factor["product"]["@id"]
                        flow = factor["exchange"]["flow"]["@id"]
                    except KeyError:
                        print(f"Skipping malformed CAUSAL_ALLOCATION in process {process_id}")
                        continue
                    if product not in allocation_dict["CAUSAL_ALLOCATION"]:
                        allocation_dict["CAUSAL_ALLOCATION"][product] = {}
                    allocation_dict["CAUSAL_ALLOCATION"][product][flow] = factor["value"]
                else:
                    product = factor.get("product", {}).get("@id")
                    if product:
                        allocation_dict[allocation_type][product] = factor["value"]
        except Exception as e:
            print(f"Error processing allocation factors in process {process_id}: {e}")
            continue
        default_method = process.get("defaultAllocationMethod")
        if default_method and default_method not in allocation_dict:
            print(f"❌ Unallocatable process: {process_id}")
            print(f"   Name: {process.get('name', 'Unnamed')}")
            print(f"   Default method: {default_method}")
            print(f"   Available methods: {list(allocation_dict.keys())}")
            print("-" * 60)

##########################################
### Find issues with default providers ###
##########################################

def check_default_provider_dict(parent_id, target_id, importer):
    """
    Checks the following:
    - defaultProvider dictionary within exchange exists
    - Checks that the four target keys exist within the defaultProvider dictionary
    - Checks that the keys are populated with a string type value

    Returns error dictionary if any of these fail, otherwise returns None
    """
    process = importer.data["processes"][parent_id]
    # find match between target id and flow id
    exchange = next(ex for ex in process["exchanges"] if ex.get("flow", {}).get("@id") == target_id)
    default_provider = exchange.get("defaultProvider", {})
    if not all(isinstance(default_provider.get(k), str) and default_provider.get(k) for k in ["@id", "name", "category", "flowType"]):
        flow = exchange.get("flow", {})
        return {
            "parentProcessID": parent_id,
            "targetID": target_id,
            "targetName": flow.get("name"),
            "targetCat": flow.get("category"),
            "targetFT": flow.get("flowType")
        }
    return None

def check_provider_exists(parent_id, target_id, importer):
    """
    Checks that there is a matching process in the importer for the target exchange being checked.
    '@id' is pulled from the defaultProvider dictionary of the target exchange.
    The dictionary importer.data['processes'] is searched for the process based on '@id'
    If no match, the data is recorded in the error dictionary and returned
    If a match, is found None is returned
    """
    process = importer.data["processes"][parent_id]
    exchange = next(ex for ex in process["exchanges"] if ex.get("flow", {}).get("@id") == target_id)
    default_provider = exchange.get("defaultProvider", {})
    if default_provider.get("@id") not in importer.data.get("processes", {}):
        return {
            "targetPrvID": default_provider.get("@id"),
            "targetPrvName": default_provider.get("name"),
            "targetPrvCat": default_provider.get("category")
        }
    return None


def provider_lacks_target_exchange(parent_id, target_id, importer):
    """
    Checks if the provider has an exchange matching the target flow.
    Returns an error dictionary if no matching exchange is found, otherwise None.
    """
    process = importer.data["processes"][parent_id]
    exchange = next(ex for ex in process["exchanges"] if ex.get("flow", {}).get("@id") == target_id)
    default_provider = exchange["defaultProvider"]
    found_provider = importer.data["processes"].get(default_provider["@id"])

    for found_exch in found_provider.get("exchanges", []):
        found_flow = found_exch.get("flow", {})
        if found_flow.get("@id") == target_id:
            return None  # Match found

    return {
        "parentProcessID": parent_id,
        "targetID": target_id,
        "foundPrvID": found_provider.get("@id")
    }

def target_exchange_provider_output(parent_id, target_id, importer):
    """
    Checks if the matching exchange in the provider is incorrectly marked as only input.
    Some processes recycle materials and will have an exchange as both an input and an output
    Returns an error dictionary if the matching exchange is an input, otherwise None.
    """
    process = importer.data["processes"][parent_id]
    exchange = next(ex for ex in process["exchanges"] if ex.get("flow", {}).get("@id") == target_id)
    default_provider = exchange["defaultProvider"]
    found_provider = importer.data["processes"].get(default_provider["@id"])

    for found_exch in found_provider.get("exchanges", []):
        found_flow = found_exch.get("flow", {})
        if found_flow.get("@id") == target_id:
            if found_exch.get("input", False):
                return {
                    "parentProcessID": parent_id,
                    "targetID": target_id,
                    "foundPrvID": found_provider.get("@id"),
                    "foundPrvExchID": found_flow.get("@id")
                }
            break

    return None

def write_provider_errors(error_dicts, output_path):
    """
    Opens an existing Excel workbook, retains only the 'Documentation' sheet,
    deletes all other sheets, and adds new sheets for each error dictionary provided.
    Each new sheet will have the top row frozen to keep column headers visible.

    Parameters:
    - error_dicts (dict): Dictionary where keys are sheet names and values are lists of dictionaries (rows).
    - output_path (str): Path to the existing Excel file to modify.

    Returns:
    - None
    """
    # Ensure the file exists
    if not os.path.exists(output_path):
        raise FileNotFoundError(f"The file {output_path} does not exist.")

    # Load the workbook
    wb = openpyxl.load_workbook(output_path)

    # Retain only the 'Documentation' sheet
    for sheet in wb.sheetnames:
        if sheet != 'Documentation':
            std = wb[sheet]
            wb.remove(std)

    # Add new sheets for each error dictionary
    for sheet_name, records in error_dicts.items():
        # Create DataFrame from list of dictionaries
        df = pd.DataFrame(records)

        # Add new sheet
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit

        # Write DataFrame to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Freeze top row
        ws.freeze_panes = "A2"

    # Save the workbook
    wb.save(output_path)

def check_default_providers(importer, output_path, debug=False):
    """
    Error checking function. Calls helper functions to identify issues with 'defaultProvider' dictionaries.
    """
    error_dicts = {
        "issueWithFlowPrvMetadata": [],
        "noMatchPrvToExc": [],
        "noMatchExcInFoundPrv": [],
        "matchExcFromPrvIsInput": []
    }

    total_processes = 0
    total_exchanges_checked = 0
    exch_not_dict = 0
    skipped_exchanges = 0
    malformed_flows = 0
    malformed_providers = 0

    for parentProcessID, process in importer.data.get("processes", {}).items():
        total_processes += 1
        for exch in process.get("exchanges", []):
            if not isinstance(exch, dict):
                exch_not_dict += 1
                continue
            if not exch.get("input", False):
                skipped_exchanges += 1
                continue
            flow = exch.get("flow", {})
            if not isinstance(flow, dict):
                malformed_flows += 1
                continue
            if flow.get("flowType") != "PRODUCT_FLOW":
                continue

            total_exchanges_checked += 1
            targetID = flow.get("@id")

            error = check_default_provider_dict(parentProcessID, targetID, importer)
            if error:
                malformed_providers += 1
                error_dicts["issueWithFlowPrvMetadata"].append(error)
                continue

            error = check_provider_exists(parentProcessID, targetID, importer)
            if error:
                error_dicts["noMatchPrvToExc"].append(error)
                continue

            errors = provider_lacks_target_exchange(parentProcessID, targetID, importer)
            if error:
                error_dicts["noMatchExcInFoundPrv"].append(error)
                continue

            errors = target_exchange_provider_output(parentProcessID, targetID, importer)
            if error:
                error_dicts["matchExcFromPrvIsInput"].append(error)
                continue

    if debug:
        print("🔍 Debug Summary:")
        print(f"Total processes checked: {total_processes}")
        print(f"Total exchanges checked: {total_exchanges_checked}")
        print(f"Total exchanges that are not dictionaries: {exch_not_dict}")
        print(f"Skipped exchanges (not input): {skipped_exchanges}")
        print(f"Malformed flow entries: {malformed_flows}")
        print(f"Malformed defaultProvider entries: {malformed_providers}")
        print("✅ Debug summary complete.")

    write_provider_errors(error_dicts, output_path)

########################################################################################################################
## Methods for working with unlinked edges
########################################################################################################################

def clean_all_locations(jsonld):
    """
    Clean and standardize 'location' fields in a JSONLDImporter object.

    This function addresses a specific issue encountered when exporting data to Excel
    using the `xlsxwriter` library, where a `TypeError: unhashable type: 'dict'` occurs.
    The root cause of this error is that some entries in the dataset contain a 'location'
    field that is either:
        - A dictionary (which is unhashable and cannot be used in Excel shared string tables),
        - `None`, or
        - Missing entirely.

    These invalid 'location' values are passed to `sheet.write_string()` in the
    `write_lci_matching()` function, which expects a string. When a non-string value
    (dict or NoneType) is passed, `xlsxwriter` fails when trying to store it in its
    internal shared string table.

    This function resolves the issue by:
        - Iterating over all entries in `uslci.data` and `uslci.products`,
        - Checking each dictionary for the presence and type of the 'location' field,
        - Replacing any non-string, missing, or None 'location' values with the string
          "no location",
        - Recursively checking all nested 'exchanges' within each dataset entry,
        - Logging each fix with the process ID and original value for traceability.

    Parameters
    ----------
    uslci : JSONLDImporter
        An instance of the JSONLDImporter class containing `.data` and `.products`
        attributes, each of which is a list of dictionaries representing LCI data.

    Returns
    -------
    None
        The function modifies the `uslci` object in-place and prints a summary of
        the changes made.
    """
    count_fixed = 0
    def clean_entry(entry, context=""):
        nonlocal count_fixed
        location = entry.get("location")
        if location is None or not isinstance(location, str):
            process_id = entry.get("id") or entry.get("code") or "(unknown ID)"
            #print(f"[{context}] FIXING Process ID: {process_id}")
            #print(f"Original location type: {type(location).__name__}")
            #print(f"Original location content: {location}")
            #print("-" * 40)
            entry["location"] = "no location"
            count_fixed += 1

    for entry in jsonld.data:
        clean_entry(entry, "DATA")
        for exc in entry.get("exchanges", []):
            clean_entry(exc, "EXCHANGE")

    for product in jsonld.products:
        clean_entry(product, "PRODUCT")

    print(f"✅ Total entries' location fixed: {count_fixed}")

def write_unlinked_flows_to_excel(importer, output_directory):
    """
    Identify and export unlinked flows from a Brightway25 JSONLDImporter object to an Excel file.

    This function:
    - Identifies unlinked exchanges using Brightway's logic.
    - Uses `activity_hash()` to ensure consistent uniqueness with Brightway's internal statistics.
    - Tracks and exports:
        1. Unique unlinked exchanges.
        2. Processes that contain unlinked exchanges.
        3. Unique processes with at least one unlinked exchange.

    Parameters:
    ----------
    importer : JSONLDImporter
        A Brightway25 importer object containing `.data` with datasets and exchanges.
    output_directory : str
        Path to the directory where the Excel file will be saved.

    Output:
    -------
    An Excel file named `unlinked_flows.xlsx` with three sheets:
        - "unique_unlinked_exc"
        - "process_with_unlinked_exc"
        - "unique_process_unlinked_exc"
    """
    # Define activity_hash inline (or import from bw2data.utils if available)
    def activity_hash(data: dict, fields: Optional[List[str]] = None, case_insensitive: bool = True) -> str:
        default_fields = ["name", "unit", "location", "type", "categories", "code"]
        lower = lambda x: x.lower() if case_insensitive and isinstance(x, str) else str(x)

        def get_value(obj, field):
            value = obj.get(field)
            if isinstance(value, (list, tuple)):
                return lower("".join(map(str, value)))
            return lower(value or "")

        fields = fields or default_fields
        string = "".join([get_value(data, field) for field in fields])
        return hashlib.md5(string.encode("utf-8")).hexdigest()

    # Prepare containers
    unique_unlinked_set = set()
    unique_unlinked_data = []
    process_with_unlinked = []
    unique_process_set = set()
    unique_process_data = []

    for ds in importer.data:
        ds_type = ds.get("type")
        ds_code = ds.get("code", "No code")
        ds_name = ds.get("name", "No name")

        has_unlinked = False

        for exc in ds.get("exchanges", []):
            if not exc.get("input") and not (ds_type == "multifunctional" and exc.get("functional")):
                exc_hash = activity_hash(exc)

                # ✅ Use hash to determine uniqueness
                if exc_hash not in unique_unlinked_set:
                    unique_unlinked_set.add(exc_hash)
                    unique_unlinked_data.append({
                        "hash": exc_hash,
                        "type": exc.get("type", "unknown"),
                        "code": exc.get("code", "No code"),
                        "name": exc.get("name", "No name"),
                        "unit": exc.get("unit", ""),
                        "location": exc.get("location", ""),
                        "categories": exc.get("categories", "")
                    })

                process_with_unlinked.append({
                    "process_code": ds_code,
                    "process_name": ds_name,
                    "unlinked_exchange_code": exc.get("code", "No code")
                })

                has_unlinked = True

        if has_unlinked and ds_code not in unique_process_set:
            unique_process_set.add(ds_code)
            unique_process_data.append({
                "code": ds_code,
                "name": ds_name
            })

    # Create DataFrames
    df_unique_unlinked = pd.DataFrame(unique_unlinked_data)
    df_process_with_unlinked = pd.DataFrame(process_with_unlinked)
    df_unique_process_unlinked = pd.DataFrame(unique_process_data)

    # Write to Excel
    output_path = os.path.join(output_directory, "unlinked_flows.xlsx")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_unique_unlinked.to_excel(writer, sheet_name="unique_unlinked_exc", index=False)
        df_process_with_unlinked.to_excel(writer, sheet_name="process_with_unlinked_exc", index=False)
        df_unique_process_unlinked.to_excel(writer, sheet_name="unique_process_unlinked_exc", index=False)

    print(f"Excel file saved to: {output_path}")